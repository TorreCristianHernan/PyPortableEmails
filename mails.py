import win32com.client
import openpyxl
from datetime import datetime, timedelta
import pytz

from utils import get_rows_to_append, build_folder_dict  # Assuming you have a utility function to extract data from email body

excel_name = "template1.xlsx"

def get_first_empty_row(worksheet):
    for row in worksheet.iter_rows():
        if not row[0].value:  # Check the first cell (column A)
            return row[0].row
    # If all rows are filled, return the next row
    return worksheet.max_row + 1

def format_date(start_date, end_date):
    start_date_input = start_date
    if start_date_input:
        start_date_input = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=pytz.UTC)

    else:
        start_date_input = pytz.utc.localize(datetime.utcnow() - timedelta(days=30))

    end_date_input = end_date
    if end_date_input:
        end_date_input = datetime.combine(end_date, datetime.min.time()).replace(tzinfo=pytz.UTC)
    else:
        end_date_input = pytz.utc.localize(datetime.utcnow() + timedelta(days=1))
    return start_date_input, end_date_input

def connect_outlook():
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.GetDefaultFolder(6)  # 6 corresponds to the Inbox folder
    return inbox

def get_inbox_folders():
    inbox = connect_outlook()
    folder_dic = build_folder_dict(inbox)
    return folder_dic

def connect_excel():
    wb = openpyxl.load_workbook(excel_name)
    ws = wb.worksheets[1]
    return wb, ws

def save_data_to_excel(start_date, end_date, subject_phrase, start_row, folder, wb, ws):
    for email in folder.Items:
        received_time = email.ReceivedTime
        # Check if the email subject contains the specified phrase and falls within the date range
        if subject_phrase.lower() in email.Subject.lower() and start_date <= received_time <= end_date:
            # Extract relevant information from the email body
            body = email.Body
            
            # Extract rows to append from email body
            rows_to_append = get_rows_to_append(email.Subject, body)
            
            # Append the extracted information to the Excel worksheet
            column_indices = {
                'Periodo': 1,
                'Nro_Solicitud': 2,
                'SDATool': 3,
                'Creador_pedido': 4,
                'MVP': 5,
                'Horas_totales': 6,
                'Perfil': 7,
                'Unidades': 8
            }
            for profile_data in rows_to_append:
                for column_name, cell_value in zip(column_indices.keys(), profile_data):
                    column_index = column_indices[column_name]
                    ws.cell(row=start_row, column=column_index, value=cell_value)
                start_row += 1

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_filename = f'output_{timestamp}.xlsx'
    
    # Save the workbook
    wb.save(output_filename)

    print("Tarea finalizada exitosamente")
            
def leer_correos_outlook_y_guardar_en_excel(start_date_input, end_date_input, inbox):
    try:
        start_date, end_date = format_date(start_date_input, end_date_input)
        
        subject_phrase = "asignacion de pedido"
                
        wb, ws = connect_excel()
        start_row = get_first_empty_row(ws)
                
        save_data_to_excel(start_date, end_date, subject_phrase, start_row, inbox, wb, ws)
    except Exception as e:
        print(f"Error al procesar los correos: {str(e)}")

# Ejecutar la función
# leer_correos_outlook_y_guardar_en_excel()